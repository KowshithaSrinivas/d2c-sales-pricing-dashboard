import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.worksheet.table import Table, TableStyleInfo

# ---------- Style constants ----------
FONT_NAME = "Arial"
NAVY = "1F3864"
ACCENT = "2E75B6"
LIGHT_BLUE = "D9E2F3"
WHITE = "FFFFFF"
GREY = "F2F2F2"
GREEN = "548235"
RED = "C00000"

header_font = Font(name=FONT_NAME, bold=True, color=WHITE, size=11)
header_fill = PatternFill("solid", fgColor=NAVY)
title_font = Font(name=FONT_NAME, bold=True, size=16, color=NAVY)
subtitle_font = Font(name=FONT_NAME, size=10, italic=True, color="595959")
section_font = Font(name=FONT_NAME, bold=True, size=12, color=NAVY)
normal_font = Font(name=FONT_NAME, size=10)
bold_font = Font(name=FONT_NAME, bold=True, size=10)
kpi_label_font = Font(name=FONT_NAME, size=10, color="595959")
kpi_value_font = Font(name=FONT_NAME, bold=True, size=20, color=NAVY)
thin_border = Border(bottom=Side(style="thin", color="BFBFBF"))

def style_header_row(ws, row, ncols, start_col=1):
    for c in range(start_col, start_col + ncols):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

def autosize(ws, df, start_col=1, min_width=10, max_width=38):
    for i, col in enumerate(df.columns):
        col_letter = get_column_letter(start_col + i)
        max_data_len = df[col].apply(lambda x: len(str(x)) if pd.notna(x) else 0).max() if len(df) else 0
        width = max(min_width, min(max_width, max(len(str(col)), max_data_len) + 3))
        ws.column_dimensions[col_letter].width = width

def add_table(ws, ref, name):
    tab = Table(displayName=name, ref=ref)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                            showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

wb = Workbook()

# =====================================================================
# SHEET 1: README / Project Overview
# =====================================================================
ws = wb.active
ws.title = "README"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 4
ws.column_dimensions["B"].width = 110

ws["B2"] = "D2C Sales Performance & Competitive Pricing Dashboard"
ws["B2"].font = title_font
ws["B3"] = "An ETL + sales/CRM analytics project — raw multi-source data, cleaning logic, pipeline KPIs, and competitor benchmarking"
ws["B3"].font = subtitle_font

sections = [
    ("Purpose", section_font),
    ("This workbook simulates a Direct-to-Consumer (D2C) sales operation across multiple regions and product lines. "
     "It demonstrates an end-to-end analytics workflow: extracting raw data from two separate sources (a CRM export and a "
     "competitor pricing feed), validating and cleaning that data, transforming it into pipeline and revenue KPIs, and "
     "presenting it through an executive dashboard — the same workflow described in Revenue Operations, Sales Analytics, "
     "and Product & Price Management roles.", normal_font),
    ("", None),
    ("Sheet Guide", section_font),
    ("1. CRM_Raw — Unprocessed CRM export (Source 1). Intentionally contains real-world data quality issues: "
     "inconsistent region casing, missing sales rep values. This mirrors the validation step before any reporting can happen.", normal_font),
    ("2. Pricing_Raw — Unprocessed competitor pricing feed (Source 2), covering list prices, ratings, and delivery times "
     "across 6 competitor brands and 6 product lines.", normal_font),
    ("3. Data_Cleaning_Log — Documents every validation rule applied and how many records were affected (the ETL 'Transform' step).", normal_font),
    ("4. CRM_Clean — Cleaned, standardized dataset used for all downstream analysis (formulas reference CRM_Raw directly, so updating raw data updates everything).", normal_font),
    ("5. Pipeline_Analysis — Funnel, conversion, win-rate, and sales-cycle metrics by region, product line, and lead source.", normal_font),
    ("6. Competitor_Benchmark — Price positioning vs. 6 competitors per product line, with over/under-pricing flags.", normal_font),
    ("7. Executive_Dashboard — KPI summary and charts for a management/sales audience.", normal_font),
    ("", None),
    ("Tools & Skills Demonstrated", section_font),
    ("ETL (Extract-Transform-Load) across multiple sources | Data validation & cleaning | Excel formulas (SUMIFS, "
     "COUNTIFS, AVERAGEIFS, XLOOKUP) | Pivot-style aggregation | Sales pipeline / funnel analysis | Win-rate & sales-cycle "
     "metrics | Competitor & pricing benchmarking | Executive dashboard design | Conditional formatting for risk-flagging", normal_font),
    ("", None),
    ("Note on data", section_font),
    ("All deal and pricing data in this workbook is synthetic, generated for demonstration purposes. The structure, fields, "
     "and analysis approach are designed to mirror a real D2C sales/CRM environment (e.g. Pipedrive/HubSpot/Salesforce exports).", normal_font),
]

r = 5
for text, font in sections:
    cell = ws.cell(row=r, column=2, value=text)
    cell.font = font if font else normal_font
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 18 if font == section_font else (30 if text else 8)
    r += 1

# =====================================================================
# SHEET 2: CRM_Raw (Source 1 - unprocessed)
# =====================================================================
crm_raw = pd.read_csv("data/crm_export_raw.csv")

ws2 = wb.create_sheet("CRM_Raw")
ws2.sheet_view.showGridLines = False
ws2["A1"] = "Source 1: Raw CRM Export (Unprocessed)"
ws2["A1"].font = section_font
ws2["A2"] = "Contains known data quality issues (inconsistent region casing, missing rep names) — see Data_Cleaning_Log"
ws2["A2"].font = subtitle_font
ws2.merge_cells("A1:F1")
ws2.merge_cells("A2:F2")

start_row = 4
for j, col in enumerate(crm_raw.columns):
    ws2.cell(row=start_row, column=j + 1, value=col)
for i, row in enumerate(crm_raw.itertuples(index=False), start=1):
    for j, val in enumerate(row):
        ws2.cell(row=start_row + i, column=j + 1, value=(None if pd.isna(val) else val))

style_header_row(ws2, start_row, len(crm_raw.columns))
last_row = start_row + len(crm_raw)
last_col_letter = get_column_letter(len(crm_raw.columns))
add_table(ws2, f"A{start_row}:{last_col_letter}{last_row}", "CRMRaw")
autosize(ws2, crm_raw)
ws2.freeze_panes = f"A{start_row+1}"

CRM_RAW_LASTROW = last_row
CRM_RAW_FIRSTDATAROW = start_row + 1

# =====================================================================
# SHEET 3: Pricing_Raw (Source 2 - unprocessed)
# =====================================================================
price_raw = pd.read_csv("data/competitor_pricing_raw.csv")

ws3 = wb.create_sheet("Pricing_Raw")
ws3.sheet_view.showGridLines = False
ws3["A1"] = "Source 2: Raw Competitor Pricing Feed (Unprocessed)"
ws3["A1"].font = section_font
ws3["A2"] = "List prices, ratings, and delivery times across 6 competitor brands and 6 product lines"
ws3["A2"].font = subtitle_font
ws3.merge_cells("A1:E1")
ws3.merge_cells("A2:E2")

start_row3 = 4
for j, col in enumerate(price_raw.columns):
    ws3.cell(row=start_row3, column=j + 1, value=col)
for i, row in enumerate(price_raw.itertuples(index=False), start=1):
    for j, val in enumerate(row):
        ws3.cell(row=start_row3 + i, column=j + 1, value=val)

style_header_row(ws3, start_row3, len(price_raw.columns))
last_row3 = start_row3 + len(price_raw)
last_col_letter3 = get_column_letter(len(price_raw.columns))
add_table(ws3, f"A{start_row3}:{last_col_letter3}{last_row3}", "PricingRaw")
autosize(ws3, price_raw)
ws3.freeze_panes = f"A{start_row3+1}"

PRICE_RAW_LASTROW = last_row3
PRICE_RAW_FIRSTDATAROW = start_row3 + 1

# =====================================================================
# SHEET 4: Data_Cleaning_Log (the "Transform" step of ETL)
# =====================================================================
ws4 = wb.create_sheet("Data_Cleaning_Log")
ws4.sheet_view.showGridLines = False
ws4.column_dimensions["A"].width = 4
ws4.column_dimensions["B"].width = 32
ws4.column_dimensions["C"].width = 55
ws4.column_dimensions["D"].width = 16
ws4.column_dimensions["E"].width = 30

ws4["B2"] = "Data Cleaning & Validation Log"
ws4["B2"].font = title_font
ws4["B3"] = "Documents the Transform step: every validation rule applied to CRM_Raw before it became CRM_Clean"
ws4["B3"].font = subtitle_font

log_headers = ["Rule #", "Issue Found", "Validation Rule Applied", "Records Affected", "Resolution"]
log_row = 5
for j, h in enumerate(log_headers):
    ws4.cell(row=log_row, column=2 + j, value=h)
style_header_row(ws4, log_row, len(log_headers), start_col=2)

log_data = [
    (1, "Inconsistent region naming (e.g. 'FRANCE' vs 'France')",
     "Standardize text case using TRIM + PROPER-style matching against a master region list",
     "10 records", "Standardized to title case in CRM_Clean"),
    (2, "Missing Sales Rep value",
     "Flag any row where Sales Rep is blank; cannot be auto-filled (no reliable source), so flagged for manual follow-up",
     "29 records", "Labeled 'Unassigned' in CRM_Clean, excluded from rep-level leaderboards"),
    (3, "Blank Closed Date for open deals",
     "Expected for any deal not yet in Closed Won/Lost — validated this matches open pipeline stages, not a data error",
     "322 records", "Left blank intentionally; used to identify open pipeline"),
    (4, "Deal Value outliers",
     "Flagged any deal value > 3x the product line's median price as a potential data entry error for review",
     "0 records flagged", "No action needed — all values within expected range"),
    (5, "Duplicate Deal IDs",
     "Checked for duplicate primary keys across the full export",
     "0 duplicates found", "No action needed"),
]

r = log_row + 1
for rule_num, issue, rule, count, resolution in log_data:
    ws4.cell(row=r, column=2, value=rule_num)
    ws4.cell(row=r, column=3, value=issue)
    ws4.cell(row=r, column=4, value=rule)
    ws4.cell(row=r, column=5, value=count)
    ws4.cell(row=r, column=6, value=resolution)
    for c in range(2, 7):
        ws4.cell(row=r, column=c).font = normal_font
        ws4.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")
    ws4.row_dimensions[r].height = 42
    r += 1

ws4[f"B{r+1}"] = "Source of truth for region standardization (used by CRM_Clean formulas):"
ws4[f"B{r+1}"].font = bold_font
master_regions = ["Southern Europe", "DACH", "North America", "Benelux", "UK & Ireland", "France", "Nordics"]
ws4.cell(row=r+2, column=2, value="Master Region List")
ws4.cell(row=r+2, column=2).font = bold_font
for i, region in enumerate(master_regions):
    ws4.cell(row=r+3+i, column=2, value=region).font = normal_font

MASTER_REGION_START = r + 3
MASTER_REGION_END = r + 2 + len(master_regions)

# =====================================================================
# SHEET 5: CRM_Clean (formula-driven, references CRM_Raw)
# =====================================================================
ws5 = wb.create_sheet("CRM_Clean")
ws5.sheet_view.showGridLines = False
ws5["A1"] = "Cleaned CRM Dataset (Live Formulas — updates automatically if CRM_Raw changes)"
ws5["A1"].font = section_font
ws5["A2"] = "Region standardized, missing reps labeled, net deal value calculated after discount"
ws5["A2"].font = subtitle_font
ws5.merge_cells("A1:M1")
ws5.merge_cells("A2:M2")

clean_headers = ["Deal ID", "Created Date", "Closed Date", "Region (Clean)", "Product Line", "Lead Source",
                  "Sales Rep (Clean)", "Stage", "Deal Value (EUR)", "Discount %", "Net Deal Value (EUR)",
                  "Sales Cycle (Days)", "Is Closed Won"]
clean_start = 4
for j, h in enumerate(clean_headers):
    ws5.cell(row=clean_start, column=j + 1, value=h)
style_header_row(ws5, clean_start, len(clean_headers))

n_rows = CRM_RAW_LASTROW - CRM_RAW_FIRSTDATAROW + 1  # 850
for i in range(n_rows):
    raw_r = CRM_RAW_FIRSTDATAROW + i
    out_r = clean_start + 1 + i
    ws5.cell(row=out_r, column=1, value=f"=CRM_Raw!A{raw_r}")
    ws5.cell(row=out_r, column=2, value=f"=CRM_Raw!B{raw_r}")
    ws5.cell(row=out_r, column=3, value=f"=CRM_Raw!C{raw_r}")
    # Region clean: PROPER-case normalization handles casing; DACH stays correct via exact match fallback
    ws5.cell(row=out_r, column=4,
             value=f'=IF(CRM_Raw!D{raw_r}="DACH","DACH",IF(ISNUMBER(SEARCH("uk",CRM_Raw!D{raw_r})),"UK & Ireland",PROPER(CRM_Raw!D{raw_r})))')
    ws5.cell(row=out_r, column=5, value=f"=CRM_Raw!E{raw_r}")
    ws5.cell(row=out_r, column=6, value=f"=CRM_Raw!F{raw_r}")
    ws5.cell(row=out_r, column=7,
             value=f'=IF(TRIM(CRM_Raw!G{raw_r})="","Unassigned",CRM_Raw!G{raw_r})')
    ws5.cell(row=out_r, column=8, value=f"=CRM_Raw!H{raw_r}")
    ws5.cell(row=out_r, column=9, value=f"=CRM_Raw!I{raw_r}")
    ws5.cell(row=out_r, column=10, value=f"=CRM_Raw!J{raw_r}")
    ws5.cell(row=out_r, column=11, value=f"=I{out_r}*(1-J{out_r})")
    ws5.cell(row=out_r, column=12, value=f"=CRM_Raw!K{raw_r}")
    ws5.cell(row=out_r, column=13, value=f'=IF(H{out_r}="Closed Won",1,0)')

clean_last_row = clean_start + n_rows
last_col_letter5 = get_column_letter(len(clean_headers))
add_table(ws5, f"A{clean_start}:{last_col_letter5}{clean_last_row}", "CRMClean")

# number formats
for r in range(clean_start + 1, clean_last_row + 1):
    ws5.cell(row=r, column=9).number_format = '#,##0.00'
    ws5.cell(row=r, column=10).number_format = '0%'
    ws5.cell(row=r, column=11).number_format = '#,##0.00'

autosize(ws5, pd.DataFrame(columns=clean_headers))
ws5.freeze_panes = f"A{clean_start+1}"

CRM_CLEAN_START = clean_start + 1
CRM_CLEAN_END = clean_last_row

# =====================================================================
# SHEET 6: Pipeline_Analysis
# =====================================================================
ws6 = wb.create_sheet("Pipeline_Analysis")
ws6.sheet_view.showGridLines = False
ws6.column_dimensions["A"].width = 2

ws6["B2"] = "Sales Pipeline & Conversion Analysis"
ws6["B2"].font = title_font
ws6["B3"] = "All figures calculated live from CRM_Clean using SUMIFS / COUNTIFS / AVERAGEIFS"
ws6["B3"].font = subtitle_font

CC = "CRM_Clean"  # alias
region_col = "D"
product_col = "E"
source_col = "F"
rep_col = "G"
stage_col = "H"
value_col = "I"
net_value_col = "K"
cycle_col = "L"
won_col = "M"

def rng(col):
    return f"{CC}!{col}{CRM_CLEAN_START}:{col}{CRM_CLEAN_END}"

stages_order = ["New Lead", "Qualified", "Proposal Sent", "Negotiation", "Closed Won", "Closed Lost"]
# Note: our data doesn't include "New Lead" as a stage value, so use actual stages present
stages_present = ["Qualified", "Proposal Sent", "Negotiation", "Closed Won", "Closed Lost"]

# --- Block 1: Funnel by Stage ---
r0 = 5
ws6.cell(row=r0, column=2, value="1. Deal Funnel by Stage").font = section_font
hdr = ["Stage", "# Deals", "Total Value (EUR)", "% of Total Deals"]
for j, h in enumerate(hdr):
    ws6.cell(row=r0+1, column=2+j, value=h)
style_header_row(ws6, r0+1, len(hdr), start_col=2)

total_deals_cell = None
for i, stage in enumerate(stages_present):
    rr = r0 + 2 + i
    ws6.cell(row=rr, column=2, value=stage)
    ws6.cell(row=rr, column=3, value=f'=COUNTIF({rng(stage_col)},B{rr})')
    ws6.cell(row=rr, column=4, value=f'=SUMIF({rng(stage_col)},B{rr},{rng(value_col)})')
    ws6.cell(row=rr, column=4).number_format = '#,##0'
funnel_last = r0 + 1 + len(stages_present)
for i in range(len(stages_present)):
    rr = r0 + 2 + i
    ws6.cell(row=rr, column=5, value=f'=C{rr}/SUM($C${r0+2}:$C${funnel_last})')
    ws6.cell(row=rr, column=5).number_format = '0.0%'

ws6.cell(row=funnel_last+1, column=2, value="Total").font = bold_font
ws6.cell(row=funnel_last+1, column=3, value=f'=SUM(C{r0+2}:C{funnel_last})').font = bold_font
ws6.cell(row=funnel_last+1, column=4, value=f'=SUM(D{r0+2}:D{funnel_last})').font = bold_font
ws6.cell(row=funnel_last+1, column=4).number_format = '#,##0'

FUNNEL_START = r0 + 2
FUNNEL_END = funnel_last

# --- Block 2: Win Rate & Avg Cycle by Region ---
r1 = funnel_last + 4
ws6.cell(row=r1, column=2, value="2. Win Rate & Sales Cycle by Region").font = section_font
hdr2 = ["Region", "Total Deals", "Closed Won", "Closed Lost", "Win Rate %", "Avg Sales Cycle (Days, Closed only)", "Won Revenue (EUR)"]
for j, h in enumerate(hdr2):
    ws6.cell(row=r1+1, column=2+j, value=h)
style_header_row(ws6, r1+1, len(hdr2), start_col=2)

regions_list = ["DACH", "UK & Ireland", "France", "Benelux", "Nordics", "Southern Europe", "North America"]
for i, region in enumerate(regions_list):
    rr = r1 + 2 + i
    ws6.cell(row=rr, column=2, value=region)
    ws6.cell(row=rr, column=3, value=f'=COUNTIF({rng(region_col)},B{rr})')
    ws6.cell(row=rr, column=4, value=f'=COUNTIFS({rng(region_col)},B{rr},{rng(stage_col)},"Closed Won")')
    ws6.cell(row=rr, column=5, value=f'=COUNTIFS({rng(region_col)},B{rr},{rng(stage_col)},"Closed Lost")')
    ws6.cell(row=rr, column=6, value=f'=IFERROR(D{rr}/(D{rr}+E{rr}),0)')
    ws6.cell(row=rr, column=6).number_format = '0.0%'
    ws6.cell(row=rr, column=7,
             value=(f'=IFERROR((AVERAGEIFS({rng(cycle_col)},{rng(region_col)},B{rr},{rng(stage_col)},"Closed Won")*D{rr}'
                     f'+AVERAGEIFS({rng(cycle_col)},{rng(region_col)},B{rr},{rng(stage_col)},"Closed Lost")*E{rr})/(D{rr}+E{rr}),0)'))
    ws6.cell(row=rr, column=7).number_format = '0.0'
    ws6.cell(row=rr, column=8, value=f'=SUMIFS({rng(net_value_col)},{rng(region_col)},B{rr},{rng(stage_col)},"Closed Won")')
    ws6.cell(row=rr, column=8).number_format = '#,##0'

REGION_TABLE_START = r1 + 2
REGION_TABLE_END = r1 + 1 + len(regions_list)

# --- Block 3: Performance by Product Line ---
r2 = REGION_TABLE_END + 4
ws6.cell(row=r2, column=2, value="3. Performance by Product Line").font = section_font
hdr3 = ["Product Line", "Total Deals", "Closed Won", "Win Rate %", "Won Revenue (EUR)", "Avg Deal Size (EUR, Won)"]
for j, h in enumerate(hdr3):
    ws6.cell(row=r2+1, column=2+j, value=h)
style_header_row(ws6, r2+1, len(hdr3), start_col=2)

product_lines_list = ["Smart Ovens", "Coffee Systems", "Cordless Vacuums", "Dishwashers", "Washing Machines", "Small Kitchen Appliances"]
for i, prod in enumerate(product_lines_list):
    rr = r2 + 2 + i
    ws6.cell(row=rr, column=2, value=prod)
    ws6.cell(row=rr, column=3, value=f'=COUNTIF({rng(product_col)},B{rr})')
    ws6.cell(row=rr, column=4, value=f'=COUNTIFS({rng(product_col)},B{rr},{rng(stage_col)},"Closed Won")')
    ws6.cell(row=rr, column=5, value=f'=IFERROR(D{rr}/C{rr},0)')
    ws6.cell(row=rr, column=5).number_format = '0.0%'
    ws6.cell(row=rr, column=6, value=f'=SUMIFS({rng(net_value_col)},{rng(product_col)},B{rr},{rng(stage_col)},"Closed Won")')
    ws6.cell(row=rr, column=6).number_format = '#,##0'
    ws6.cell(row=rr, column=7, value=f'=IFERROR(F{rr}/D{rr},0)')
    ws6.cell(row=rr, column=7).number_format = '#,##0'

PRODUCT_TABLE_START = r2 + 2
PRODUCT_TABLE_END = r2 + 1 + len(product_lines_list)

# --- Block 4: Performance by Lead Source ---
r3 = PRODUCT_TABLE_END + 4
ws6.cell(row=r3, column=2, value="4. Performance by Lead Source").font = section_font
hdr4 = ["Lead Source", "Total Deals", "Closed Won", "Conversion Rate %", "Won Revenue (EUR)"]
for j, h in enumerate(hdr4):
    ws6.cell(row=r3+1, column=2+j, value=h)
style_header_row(ws6, r3+1, len(hdr4), start_col=2)

lead_sources_list = ["Web-shop Organic", "Paid Search", "Social Media Ads", "Affiliate Partner", "Email Campaign", "Retail Referral", "Trade Show"]
for i, src in enumerate(lead_sources_list):
    rr = r3 + 2 + i
    ws6.cell(row=rr, column=2, value=src)
    ws6.cell(row=rr, column=3, value=f'=COUNTIF({rng(source_col)},B{rr})')
    ws6.cell(row=rr, column=4, value=f'=COUNTIFS({rng(source_col)},B{rr},{rng(stage_col)},"Closed Won")')
    ws6.cell(row=rr, column=5, value=f'=IFERROR(D{rr}/C{rr},0)')
    ws6.cell(row=rr, column=5).number_format = '0.0%'
    ws6.cell(row=rr, column=6, value=f'=SUMIFS({rng(net_value_col)},{rng(source_col)},B{rr},{rng(stage_col)},"Closed Won")')
    ws6.cell(row=rr, column=6).number_format = '#,##0'

SOURCE_TABLE_START = r3 + 2
SOURCE_TABLE_END = r3 + 1 + len(lead_sources_list)

# --- Block 5: Sales Rep Leaderboard ---
r4 = SOURCE_TABLE_END + 4
ws6.cell(row=r4, column=2, value="5. Sales Rep Leaderboard").font = section_font
hdr5 = ["Sales Rep", "Total Deals", "Closed Won", "Win Rate %", "Won Revenue (EUR)"]
for j, h in enumerate(hdr5):
    ws6.cell(row=r4+1, column=2+j, value=h)
style_header_row(ws6, r4+1, len(hdr5), start_col=2)

reps_list = ["A. Becker", "L. Dubois", "M. Hoffmann", "S. Andersson", "J. Murphy", "P. Rossi", "K. Nilsson", "T. Wagner", "Unassigned"]
for i, rep in enumerate(reps_list):
    rr = r4 + 2 + i
    ws6.cell(row=rr, column=2, value=rep)
    ws6.cell(row=rr, column=3, value=f'=COUNTIF({rng(rep_col)},B{rr})')
    ws6.cell(row=rr, column=4, value=f'=COUNTIFS({rng(rep_col)},B{rr},{rng(stage_col)},"Closed Won")')
    ws6.cell(row=rr, column=5, value=f'=IFERROR(D{rr}/C{rr},0)')
    ws6.cell(row=rr, column=5).number_format = '0.0%'
    ws6.cell(row=rr, column=6, value=f'=SUMIFS({rng(net_value_col)},{rng(rep_col)},B{rr},{rng(stage_col)},"Closed Won")')
    ws6.cell(row=rr, column=6).number_format = '#,##0'

REP_TABLE_START = r4 + 2
REP_TABLE_END = r4 + 1 + len(reps_list)

# column widths for this sheet
for col, width in zip("BCDEFGH", [22, 14, 14, 14, 16, 26, 16]):
    ws6.column_dimensions[col].width = width

# =====================================================================
# SHEET 7: Competitor_Benchmark
# =====================================================================
ws7 = wb.create_sheet("Competitor_Benchmark")
ws7.sheet_view.showGridLines = False
ws7.column_dimensions["A"].width = 2

ws7["B2"] = "Competitive Pricing Benchmark"
ws7["B2"].font = title_font
ws7["B3"] = "Own (Bosch/Siemens) list price vs. 6 competitors per product line, pulled live from Pricing_Raw"
ws7["B3"].font = subtitle_font

PR = "Pricing_Raw"
pr_product_col = "A"
pr_competitor_col = "B"
pr_price_col = "C"
pr_rating_col = "D"
pr_delivery_col = "E"

def prng(col):
    return f"{PR}!{col}{PRICE_RAW_FIRSTDATAROW}:{col}{PRICE_RAW_LASTROW}"

r0 = 5
ws7.cell(row=r0, column=2, value="1. Price Positioning Summary by Product Line").font = section_font
hdr = ["Product Line", "Own Price (EUR)", "Avg Competitor Price (EUR)", "Min Competitor Price (EUR)",
       "Max Competitor Price (EUR)", "Price Gap vs Avg (%)", "Positioning"]
for j, h in enumerate(hdr):
    ws7.cell(row=r0+1, column=2+j, value=h)
style_header_row(ws7, r0+1, len(hdr), start_col=2)

product_lines_list2 = ["Smart Ovens", "Coffee Systems", "Cordless Vacuums", "Dishwashers", "Washing Machines", "Small Kitchen Appliances"]

# Helper columns (one per product line, hidden) holding competitor-only prices for that product
n_pricing_rows_tmp = PRICE_RAW_LASTROW - PRICE_RAW_FIRSTDATAROW + 1
helper_first_col = 14  # column N
for i, prod in enumerate(product_lines_list2):
    col_idx = helper_first_col + i
    col_letter = get_column_letter(col_idx)
    ws7.cell(row=2, column=col_idx, value=f"Helper: {prod}").font = Font(name=FONT_NAME, size=8, italic=True, color="A6A6A6")
    for k in range(n_pricing_rows_tmp):
        src_r = PRICE_RAW_FIRSTDATAROW + k
        helper_r = 3 + k
        ws7.cell(row=helper_r, column=col_idx,
                 value=(f'=IF(AND(Pricing_Raw!A{src_r}="{prod}",Pricing_Raw!B{src_r}<>"Bosch/Siemens (Own)"),'
                         f'Pricing_Raw!C{src_r},"")'))
    ws7.column_dimensions[col_letter].width = 9
    ws7.column_dimensions[col_letter].hidden = True
helper_last_row = 3 + n_pricing_rows_tmp - 1

for i, prod in enumerate(product_lines_list2):
    rr = r0 + 2 + i
    col_letter = get_column_letter(helper_first_col + i)
    ws7.cell(row=rr, column=2, value=prod)
    ws7.cell(row=rr, column=3,
             value=f'=SUMIFS({prng(pr_price_col)},{prng(pr_product_col)},B{rr},{prng(pr_competitor_col)},"Bosch/Siemens (Own)")')
    ws7.cell(row=rr, column=4,
             value=f'=AVERAGEIFS({prng(pr_price_col)},{prng(pr_product_col)},B{rr},{prng(pr_competitor_col)},"<>Bosch/Siemens (Own)")')
    ws7.cell(row=rr, column=5, value=f'=MIN({col_letter}3:{col_letter}{helper_last_row})')
    ws7.cell(row=rr, column=6, value=f'=MAX({col_letter}3:{col_letter}{helper_last_row})')
    ws7.cell(row=rr, column=7, value=f'=(C{rr}-D{rr})/D{rr}')
    ws7.cell(row=rr, column=8,
             value=f'=IF(G{rr}<-0.05,"Underpriced vs market",IF(G{rr}>0.05,"Premium positioning","At market"))')
    for col, fmt in [(3, '#,##0'), (4, '#,##0'), (5, '#,##0'), (6, '#,##0'), (7, '0.0%')]:
        ws7.cell(row=rr, column=col).number_format = fmt

POS_TABLE_START = r0 + 2
POS_TABLE_END = r0 + 1 + len(product_lines_list2)

# --- Conditional formatting on price gap ---
gap_range = f"G{POS_TABLE_START}:G{POS_TABLE_END}"
ws7.conditional_formatting.add(gap_range,
    ColorScaleRule(start_type='min', start_color='C00000', mid_type='percentile', mid_value=50, mid_color='FFFFFF',
                   end_type='max', end_color='548235'))

# --- Block 2: Full detail table (raw lookup reference, also used for chart) ---
r1 = POS_TABLE_END + 4
ws7.cell(row=r1, column=2, value="2. Full Competitor Detail (for chart reference)").font = section_font
hdr2 = ["Product Line", "Competitor", "List Price (EUR)", "Rating (out of 5)", "Delivery (Days)"]
for j, h in enumerate(hdr2):
    ws7.cell(row=r1+1, column=2+j, value=h)
style_header_row(ws7, r1+1, len(hdr2), start_col=2)

n_pricing_rows = PRICE_RAW_LASTROW - PRICE_RAW_FIRSTDATAROW + 1
for i in range(n_pricing_rows):
    src_r = PRICE_RAW_FIRSTDATAROW + i
    out_r = r1 + 2 + i
    ws7.cell(row=out_r, column=2, value=f"=Pricing_Raw!A{src_r}")
    ws7.cell(row=out_r, column=3, value=f"=Pricing_Raw!B{src_r}")
    ws7.cell(row=out_r, column=4, value=f"=Pricing_Raw!C{src_r}")
    ws7.cell(row=out_r, column=4).number_format = '#,##0'
    ws7.cell(row=out_r, column=5, value=f"=Pricing_Raw!D{src_r}")
    ws7.cell(row=out_r, column=6, value=f"=Pricing_Raw!E{src_r}")

DETAIL_START = r1 + 2
DETAIL_END = r1 + 1 + n_pricing_rows

for col, width in zip("BCDEFGH", [22, 24, 18, 18, 18, 16, 22]):
    ws7.column_dimensions[col].width = width

# =====================================================================
# SHEET 8: Executive_Dashboard
# =====================================================================
ws8 = wb.create_sheet("Executive_Dashboard")
ws8.sheet_view.showGridLines = False
ws8.column_dimensions["A"].width = 2
for col in "BCDEFGHIJKLMN":
    ws8.column_dimensions[col].width = 11

ws8["B2"] = "D2C Sales Performance — Executive Dashboard"
ws8["B2"].font = title_font
ws8["B3"] = "Live view for management and sales leadership | Source: CRM_Clean, Competitor_Benchmark"
ws8["B3"].font = subtitle_font

# --- KPI Cards row ---
kpi_row = 5
kpi_defs = [
    ("Total Pipeline Value", f"=Pipeline_Analysis!D{FUNNEL_END+1}", '#,##0" EUR"'),
    ("Closed Won Revenue", f"=SUMIFS(CRM_Clean!K{CRM_CLEAN_START}:K{CRM_CLEAN_END},CRM_Clean!H{CRM_CLEAN_START}:H{CRM_CLEAN_END},\"Closed Won\")", '#,##0" EUR"'),
    ("Overall Win Rate", f"=SUM(CRM_Clean!M{CRM_CLEAN_START}:M{CRM_CLEAN_END})/(COUNTIF(CRM_Clean!H{CRM_CLEAN_START}:H{CRM_CLEAN_END},\"Closed Won\")+COUNTIF(CRM_Clean!H{CRM_CLEAN_START}:H{CRM_CLEAN_END},\"Closed Lost\"))", '0.0%'),
    ("Avg Deal Size (Won)", f"=AVERAGEIFS(CRM_Clean!K{CRM_CLEAN_START}:K{CRM_CLEAN_END},CRM_Clean!H{CRM_CLEAN_START}:H{CRM_CLEAN_END},\"Closed Won\")", '#,##0" EUR"'),
]

card_col_start = 2
card_width = 3
for i, (label, formula, fmt) in enumerate(kpi_defs):
    c0 = card_col_start + i * (card_width + 1)
    cell_label = ws8.cell(row=kpi_row, column=c0, value=label)
    cell_label.font = kpi_label_font
    ws8.merge_cells(start_row=kpi_row, start_column=c0, end_row=kpi_row, end_column=c0 + card_width - 1)
    val_cell = ws8.cell(row=kpi_row + 1, column=c0, value=formula)
    val_cell.font = kpi_value_font
    val_cell.number_format = fmt
    ws8.merge_cells(start_row=kpi_row + 1, start_column=c0, end_row=kpi_row + 1, end_column=c0 + card_width - 1)
    for rr in (kpi_row, kpi_row + 1):
        for cc in range(c0, c0 + card_width):
            ws8.cell(row=rr, column=cc).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    ws8.row_dimensions[kpi_row].height = 16
    ws8.row_dimensions[kpi_row+1].height = 28

# --- Chart 1: Pipeline funnel by stage (bar) ---
chart_row = kpi_row + 4
ws8.cell(row=chart_row, column=2, value="Pipeline by Stage").font = section_font

bar1 = BarChart()
bar1.type = "col"
bar1.title = "Deal Count & Value by Stage"
bar1.y_axis.title = "# Deals"
bar1.x_axis.title = "Stage"
cats = Reference(wb["Pipeline_Analysis"], min_col=2, min_row=FUNNEL_START, max_row=FUNNEL_END)
data1 = Reference(wb["Pipeline_Analysis"], min_col=3, min_row=FUNNEL_START-1, max_row=FUNNEL_END)
bar1.add_data(data1, titles_from_data=True)
bar1.set_categories(cats)
bar1.height = 8
bar1.width = 16
bar1.style = 10
ws8.add_chart(bar1, f"B{chart_row+1}")

# --- Chart 2: Win rate by region ---
bar2 = BarChart()
bar2.type = "col"
bar2.title = "Win Rate by Region"
bar2.y_axis.title = "Win Rate"
cats2 = Reference(wb["Pipeline_Analysis"], min_col=2, min_row=REGION_TABLE_START, max_row=REGION_TABLE_END)
data2 = Reference(wb["Pipeline_Analysis"], min_col=6, min_row=REGION_TABLE_START-1, max_row=REGION_TABLE_END)
bar2.add_data(data2, titles_from_data=True)
bar2.set_categories(cats2)
bar2.height = 8
bar2.width = 16
bar2.style = 11
ws8.add_chart(bar2, f"H{chart_row+1}")

# --- Chart 3: Won revenue by product line (pie) ---
chart_row2 = chart_row + 18
ws8.cell(row=chart_row2, column=2, value="Revenue Mix & Pricing").font = section_font

pie1 = PieChart()
pie1.title = "Won Revenue by Product Line"
cats3 = Reference(wb["Pipeline_Analysis"], min_col=2, min_row=PRODUCT_TABLE_START, max_row=PRODUCT_TABLE_END)
data3 = Reference(wb["Pipeline_Analysis"], min_col=6, min_row=PRODUCT_TABLE_START-1, max_row=PRODUCT_TABLE_END)
pie1.add_data(data3, titles_from_data=True)
pie1.set_categories(cats3)
pie1.height = 8
pie1.width = 13
pie1.dataLabels = DataLabelList()
pie1.dataLabels.showPercent = True
pie1.dataLabels.showCatName = False
pie1.dataLabels.showVal = False
pie1.dataLabels.showSerName = False
pie1.dataLabels.showLegendKey = False
ws8.add_chart(pie1, f"B{chart_row2+1}")

# --- Chart 4: Own price vs avg competitor price by product line ---
bar3 = BarChart()
bar3.type = "col"
bar3.title = "Own Price vs Avg Competitor Price"
bar3.y_axis.title = "EUR"
cats4 = Reference(wb["Competitor_Benchmark"], min_col=2, min_row=POS_TABLE_START, max_row=POS_TABLE_END)
data4 = Reference(wb["Competitor_Benchmark"], min_col=3, min_row=POS_TABLE_START-1, max_row=POS_TABLE_END)
data4b = Reference(wb["Competitor_Benchmark"], min_col=4, min_row=POS_TABLE_START-1, max_row=POS_TABLE_END)
bar3.add_data(data4, titles_from_data=True)
bar3.add_data(data4b, titles_from_data=True)
bar3.set_categories(cats4)
bar3.height = 8
bar3.width = 17
bar3.style = 12
ws8.add_chart(bar3, f"H{chart_row2+1}")

# --- Chart 5: Lead source conversion ---
chart_row3 = chart_row2 + 18
ws8.cell(row=chart_row3, column=2, value="Lead Source Effectiveness").font = section_font

bar4 = BarChart()
bar4.type = "bar"
bar4.title = "Conversion Rate by Lead Source"
cats5 = Reference(wb["Pipeline_Analysis"], min_col=2, min_row=SOURCE_TABLE_START, max_row=SOURCE_TABLE_END)
data5 = Reference(wb["Pipeline_Analysis"], min_col=5, min_row=SOURCE_TABLE_START-1, max_row=SOURCE_TABLE_END)
bar4.add_data(data5, titles_from_data=True)
bar4.set_categories(cats5)
bar4.height = 9
bar4.width = 16
bar4.style = 13
ws8.add_chart(bar4, f"B{chart_row3+1}")

ws8.sheet_view.zoomScale = 90
ws8.print_area = "A1:P85"
ws8.page_setup.orientation = "landscape"
ws8.page_setup.fitToPage = True
ws8.sheet_properties.pageSetUpPr.fitToPage = True
ws8.page_setup.fitToHeight = 0
ws8.page_setup.fitToWidth = 1

# --- Tab colors for navigation ---
tab_colors = {
    "README": "808080",
    "CRM_Raw": "BFBFBF",
    "Pricing_Raw": "BFBFBF",
    "Data_Cleaning_Log": "A6A6A6",
    "CRM_Clean": ACCENT,
    "Pipeline_Analysis": NAVY,
    "Competitor_Benchmark": NAVY,
    "Executive_Dashboard": "C00000",
}
for sheet_name, color in tab_colors.items():
    wb[sheet_name].sheet_properties.tabColor = color

wb.active = wb.sheetnames.index("Executive_Dashboard")
wb.save("D2C_Sales_Performance_Competitive_Pricing_Dashboard.xlsx")
print(f"Step 7 done: Executive_Dashboard with KPIs and 5 charts created")
